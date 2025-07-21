from openpyxl import load_workbook, Workbook
from alive_progress import alive_it
from shutil import copy2
from os import listdir, path

def convert_to_int(element):
    try:
        element = int(element)
        return element
    except:
        return element


def list_column_values(filename, column):
    workbook = load_workbook(filename)
    sheet = workbook.active

    column_values = []
    for row in sheet.iter_rows():
        cell_value = row[column].value
        column_values.append(cell_value)

    return column_values[1:]


def converter(pictures_source_path, pictures_destination_path, excel_path, old_ext, new_ext, prefix, id_column,
              card_column):
    print("<*>  Please Wait...")
    ids = list_column_values(excel_path, id_column)
    ids = list(map(convert_to_int, ids))
    cards = list_column_values(excel_path, card_column)
    cards = list(map(convert_to_int, cards))

    pictures = listdir(pictures_source_path)
    bar = alive_it(pictures)
    destination_pictures = listdir(pictures_destination_path)
    no_id = []

    for picture in bar:
        try:
            if picture.startswith(prefix) and picture.endswith(old_ext):
                pure_picture_name = int(picture[len(prefix):(-(len(old_ext)))])
            else:
                continue
        except:
            continue
        if picture not in destination_pictures:
            if pure_picture_name in ids:
                ind = ids.index(pure_picture_name)
                new_name = str(cards[ind]) + new_ext

            elif pure_picture_name in cards:
                ind = cards.index(pure_picture_name)
                new_name = str(pure_picture_name) + new_ext

            else:
                no_id.append(picture)
                continue

            copy2(path.join(pictures_source_path, picture), path.join(pictures_destination_path, new_name))

        else:
            if pure_picture_name in ids:
                ind = ids.index(pure_picture_name)
            elif pure_picture_name in cards:
                ind = cards.index(pure_picture_name)
        try:
          ids.pop(ind)
          cards.pop(ind)
        except:
          pass

    # region report
    wb = Workbook()
    ws = wb.active
    if no_id != []:
        ws.cell(1, 1).value = "There is no ID for this pictures:"
        for i in range(len(no_id)):
            ws.cell(i + 2, 1).value = no_id[i]
        wb.save('excel.xlsx')
        wb.close()
        print(f'<*>  There is no ID for this pictures: {no_id}')

    if ids != []:
        ws.cell(1, 2).value = "There is no Picture for this IDs:"
        for i in range(len(ids)):
            ws.cell(i + 2, 2).value = ids[i]
        wb.save('excel.xlsx')
        wb.close()
        print(f'<*>  There is no Picture for this IDs: {ids}')
    # endregion

    print("<*>  Converting is Done.")